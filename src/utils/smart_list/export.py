from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any

from flask import Response

from utils.smart_list.config import SmartListConfig


def _get_cell_value(row: Any, key: str) -> str:
    """Extrai valor de um row (dict, objeto ORM ou dataclass)."""
    if isinstance(row, dict):
        val = row.get(key, "")
    else:
        val = getattr(row, key, "")
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y %H:%M")
    return str(val)


def export_csv(
    config: SmartListConfig,
    items: list[Any],
    visible_columns: list[str] | None = None,
) -> Response:
    """Gera resposta CSV para download."""
    cols = visible_columns or [c.key for c in config.columns]
    headers = {c.key: c.label for c in config.columns}

    output = io.StringIO()
    writer = csv.writer(output)

    # Cabeçalho
    writer.writerow([headers.get(k, k) for k in cols])

    # Dados
    for row in items:
        writer.writerow([_get_cell_value(row, k) for k in cols])

    filename = f"{config.export_filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def export_excel(
    config: SmartListConfig,
    items: list[Any],
    visible_columns: list[str] | None = None,
) -> Response:
    """Gera resposta Excel (.xlsx) para download."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return Response("openpyxl não instalado. Execute: pip install openpyxl", status=500)

    cols = visible_columns or [c.key for c in config.columns]
    headers = {c.key: c.label for c in config.columns}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = config.export_filename[:31]  # Excel limita a 31 chars

    # Estilos do cabeçalho
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    # Linha de cabeçalho
    for col_idx, key in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=headers.get(key, key))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Dados
    for row_idx, row in enumerate(items, 2):
        for col_idx, key in enumerate(cols, 1):
            ws.cell(row=row_idx, column=col_idx, value=_get_cell_value(row, key))

    # Auto-width aproximado
    for col_idx, key in enumerate(cols, 1):
        max_len = max(
            len(headers.get(key, key)),
            *(len(_get_cell_value(row, key)) for row in items[:50]),  # amostra
        ) if items else len(headers.get(key, key))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # Congela linha de cabeçalho
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{config.export_filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def export_pdf(
    config: SmartListConfig,
    items: list[Any],
    visible_columns: list[str] | None = None,
    title: str | None = None,
) -> Response:
    """
    Gera resposta PDF horizontal com cabeçalho.
    Requer reportlab: pip install reportlab
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return Response("reportlab não instalado. Execute: pip install reportlab", status=500)

    cols = visible_columns or [c.key for c in config.columns]
    headers = {c.key: c.label for c in config.columns}

    output = io.BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        output,
        pagesize=page_size,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=20 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    story = []

    # Título
    report_title = title or config.export_filename.replace("_", " ").title()
    story.append(Paragraph(
        f"<b>{report_title}</b> &nbsp;&nbsp; <font size=9 color='grey'>"
        f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        f" &nbsp;·&nbsp; {len(items)} registros</font>",
        styles["Heading2"],
    ))
    story.append(Spacer(1, 6 * mm))

    # Dados da tabela
    table_data = [[headers.get(k, k) for k in cols]]
    for row in items:
        table_data.append([_get_cell_value(row, k) for k in cols])

    # Largura proporcional
    usable_width = page_size[0] - 30 * mm
    col_width = usable_width / len(cols)
    col_widths = [col_width] * len(cols)

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0),  colors.HexColor("#2C3E50")),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0),  9),
        ("ALIGN",        (0, 0), (-1, 0),  "CENTER"),
        ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",     (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F8FA")]),
        ("GRID",         (0, 0), (-1, -1), 0.25, colors.HexColor("#DEE2E6")),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))

    story.append(table)
    doc.build(story)
    output.seek(0)

    filename = f"{config.export_filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        output.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
